from __future__ import annotations

from argparse import ArgumentParser, Namespace
from pathlib import Path

from openpyxl import Workbook, load_workbook

YEAR_FILTER = 2026
SALE_DURATION_FILTER = "3-day sale"
DEFAULT_INPUT = Path("3daysalesample.xlsx")
DEFAULT_OUTPUT = Path("3daysales_2026.xlsx")


def build_workbook(input_path: Path, output_path: Path) -> Path:
    source_workbook = load_workbook(input_path)
    source_sheet = source_workbook.active

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = source_sheet.title

    header = [cell.value for cell in next(source_sheet.iter_rows(min_row=1, max_row=1))]
    output_sheet.append(header)

    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        year = row[1]
        sale_duration = row[5]
        if year == YEAR_FILTER and sale_duration == SALE_DURATION_FILTER:
            output_sheet.append(row)

    date_format = source_sheet["D2"].number_format
    for cell in output_sheet["D"][1:]:
        cell.number_format = date_format

    for column_letter, dimension in source_sheet.column_dimensions.items():
        if dimension.width is not None:
            output_sheet.column_dimensions[column_letter].width = dimension.width

    output_workbook.save(output_path)
    return output_path


def parse_args() -> Namespace:
    parser = ArgumentParser(description="Generate a 2026 3-day sale Excel workbook.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Source workbook path")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output workbook path")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = build_workbook(args.input, args.output)
    print(f"Generated {result}")


if __name__ == "__main__":
    main()
